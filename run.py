import telebot
from uparallel import uparallel
import config
import time
from datetime import datetime
import logging
import zipfile
import os
from db_tool import DbQuery
from openpyxl import Workbook

db_query = DbQuery()
bot = telebot.TeleBot(config.token)
logging.basicConfig(filename="sample.log", level=logging.INFO)

markup = telebot.types.InlineKeyboardMarkup()
markup.row(telebot.types.InlineKeyboardButton('🙁', callback_data='1'),
           telebot.types.InlineKeyboardButton('😐', callback_data='2'),
           telebot.types.InlineKeyboardButton('😀', callback_data='3'))


def add_new_polling(chat_id, message_id, message_text):
    query = """SELECT *
        	        FROM public.polls
                    WHERE msg_id={};"""
    query_result = db_query.execute_query(query.format(message_id))
    if len(query_result.value) < 1:
        query = """INSERT INTO public.polls
                    (chat_id, msg_id,votes,sum,text,date)
	         VALUES ({}, {},0,0,'{}',to_timestamp('{}','YYYY-MM-DD HH24:MI:SS'));"""
        query_result = db_query.execute_query(query.format(chat_id, message_id, message_text, datetime.now()),
                                              is_dml=True)
        create_user_list(message_id)
        add_new_polling(chat_id, message_id, message_text)
    else:
        return query_result.value


def create_user_list(message_id):
    query = """CREATE TABLE public."{}"
            (user_id integer NOT NULL)
            WITH (
            OIDS = FALSE
            )
            TABLESPACE pg_default;
            ALTER TABLE public."{}"
            OWNER to root;"""
    query_result = db_query.execute_query(query.format(message_id, message_id), is_dml=True)
    print('CREATE USER LIST')
    print(query_result.success)


def user_is_new(message_id, user_id):
    query = """SELECT *
            	        FROM public."{}"
                        WHERE user_id={};"""
    query_result = db_query.execute_query(query.format(message_id, user_id))
    if len(query_result.value) < 1:
        query = """INSERT INTO public."{}"
                        (user_id)
    	         VALUES ({});"""
        query_result = db_query.execute_query(query.format(message_id, user_id), is_dml=True)
        return True
    else:
        return False

def add_vote(message_id, votes):
    query = """UPDATE public.polls
	            SET votes={}
	            WHERE msg_id={};"""
    query_result = db_query.execute_query(query.format(int(votes) + 1, message_id), is_dml=True)


def new_sum(message_id, new_summa):
    query = """UPDATE public.polls
    	            SET sum={}
    	            WHERE msg_id={};"""
    query_result = db_query.execute_query(query.format(new_summa, message_id), is_dml=True)


@bot.message_handler(regexp="/test")
def test(message):
    questions = ["Я понимаю цели проекта",
                 "Мне понятны границы ответственности участников проекта",
                 "Я согласен с текущими приоритетами проекта"]
    for msg in questions:
        bot.send_message(message.chat.id, msg, reply_markup=markup)
    pass


# @bot.message_handler(regexp="/all_result")
# def all_result(message):
#     workbook = Workbook()
#     for i in range(10):
#         workbook.create_sheet(title="Number"+str(i))
#     workbook.save(filename='test')
#
#
#     pass

@bot.message_handler(regexp="/radushin")
def test(message):
    working_directory = os.path.dirname(os.path.abspath(__file__))
    workbook = Workbook()
    query = """SELECT *
        	        FROM public.polls
        	        ORDER by chat_id ASC, id ASC;"""
    query_result = db_query.execute_query(query)
    temp_chat_name = ''
    # chat_names = []
    for data in query_result.value:
        # print(data)
        # print(data[1])
        # print(bot.get_chat(data[1]).title)
        if temp_chat_name == bot.get_chat(data[1]).title:
            if i % 3 == 0:
                work_sheet['A' + str(row)] = str(data[6])
                work_sheet['B' + str(row)] = str(data[4]).replace('.', ',')
                work_sheet['C' + str(row)] = str(data[3])
                i = i + 1
                continue
            elif i % 3 == 1:
                work_sheet['D' + str(row)] = str(data[4]).replace('.', ',')
                work_sheet['E' + str(row)] = str(data[3])
                i = i + 1
                continue
            elif i % 3 == 2:
                work_sheet['F' + str(row)] = str(data[4]).replace('.', ',')
                work_sheet['G' + str(row)] = str(data[3])
                i = i + 1
                row = row + 1
                continue
        else:
            # chat_names.append(bot.get_chat(data[1]).title)
            work_sheet = workbook.create_sheet(bot.get_chat(data[1]).title)
            work_sheet['B1'] = 'Question1'
            work_sheet['C1'] = 'Voters1'
            work_sheet['D1'] = 'Question2'
            work_sheet['E1'] = 'Voters2'
            work_sheet['F1'] = 'Question3'
            work_sheet['G1'] = 'Voters3'
            work_sheet['A2'] = str(data[6])
            work_sheet['B2'] = str(data[4]).replace('.', ',')
            work_sheet['C2'] = str(data[3])
            temp_chat_name = bot.get_chat(data[1]).title
            i = 1
            row = 2
    # print(chat_names)
    #
    # for name in chat_names:
    #     workbook.create_sheet(name)

    workbook.save('test.xlsx')

    z = zipfile.ZipFile('test.zip', 'w')
    z.write('test.xlsx')
    z.close()

    # test_file = open(, 'rb')
    bot.send_document(message.chat.id, open(working_directory + '/test.zip', 'rb'))
    # bot.send_document(message.chat.id,working_directory+'/test.zip')
    pass


@bot.callback_query_handler(func=lambda call: True)
def callback_inline(call):
    if call.message:
        if call.data:
            # bot.answer_callback_query(call.id, text="Done!")
            # user_id = json.loads(call.message)[0]
            print(1)
            data = add_new_polling(call.message.chat.id, call.message.message_id, call.message.text)
            try:
                votes = data[0][3]
                summa = data[0][4]
            except:
                votes = 0
                summa = 0
            print('votes ' + str(votes))
            print('summa ' + str(summa))
            if user_is_new(call.message.message_id, call.from_user.id):
                new_summa = (int(votes) * float(summa) + int(call.data)) / (int(votes) + 1)
                add_vote(call.message.message_id, votes)
                new_sum(call.message.message_id, new_summa)
                bot.answer_callback_query(call.id, text=str(call.data))
            else:
                print(2)
                bot.answer_callback_query(call.id, text='Already voted')
    pass


@bot.message_handler(regexp="/result")
def get_result(message):
    query = """SELECT *
                	        FROM public.polls
                            WHERE chat_id={}
                            ORDER BY msg_id DESC
                            LIMIT 3;"""
    query_result = db_query.execute_query(query.format(message.chat.id))
    print(query_result.value)
    msg = """Текст сообщения: {} \n
    Количество ответов: {} \n
    Среднее значение: {} \n
            """
    for data in query_result.value:
        bot.send_message(message.chat.id, msg.format(data[5], data[3], data[4]))

while True:
    # bot.polling(none_stop=True)
    try:
        bot.polling(none_stop=True)
    except Exception as e:
        logging.error(e)
        time.sleep(15)
