from db_tool import DbQuery
import zipfile
from openpyxl import Workbook
import config
import telebot

bot = telebot.TeleBot(config.token)

db_query = DbQuery()
workbook = Workbook()
query = """SELECT *
    	        FROM public.polls
    	        ORDER by chat_id ASC, id ASC;"""
query_result = db_query.execute_query(query)
temp_chat_name = ''
# chat_names = []
for data in query_result.value:
    # print(data)
    # print(data[1])
    # print(bot.get_chat(data[1]).title)
    if temp_chat_name ==bot.get_chat(data[1]).title:
        if i % 3 == 0 :
            work_sheet['A' + str(row)] = str(data[6])
            work_sheet['B' + str(row)] = str(data[4]).replace('.', ',')
            work_sheet['C' + str(row)] = str(data[3])
            i = i + 1
            continue
        elif i % 3 == 1:
            work_sheet['D' + str(row)] = str(data[4]).replace('.',',')
            work_sheet['E' + str(row)] = str(data[3])
            i= i +1
            continue
        elif i % 3 == 2:
            work_sheet['F' + str(row)] = str(data[4]).replace('.', ',')
            work_sheet['G' + str(row)] = str(data[3])
            i = i +1
            row = row +1
            continue
    else:
        # chat_names.append(bot.get_chat(data[1]).title)
        work_sheet = workbook.create_sheet(bot.get_chat(data[1]).title)
        work_sheet['B1'] = 'Question1'
        work_sheet['C1'] = 'Voters1'
        work_sheet['D1'] = 'Question2'
        work_sheet['E1'] = 'Voters2'
        work_sheet['F1'] = 'Question3'
        work_sheet['G1'] = 'Voters3'
        work_sheet['A2'] = str(data[6])
        work_sheet['B2'] = str(data[4]).replace('.',',')
        work_sheet['C2'] = str(data[3])
        temp_chat_name = bot.get_chat(data[1]).title
        i=1
        row = 2
# print(chat_names)
#
# for name in chat_names:
#     workbook.create_sheet(name)

workbook.save('test.xlsx')



z = zipfile.ZipFile('test.zip', 'w')
z.write('test.xlsx')
z.close()
